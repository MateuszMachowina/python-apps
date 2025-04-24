import os
import requests
from dotenv import load_dotenv
from openpyxl import load_workbook
from tkinter import Tk
from tkinter.filedialog import askopenfilename


def get_exchange_rate(base_currency, target_currency, api_key):
    url = f"https://v6.exchangerate-api.com/v6/{api_key}/latest/{base_currency}"
    try:
        response = requests.get(url)
        data = response.json()
        if data.get("result") == "success":
            return data["conversion_rates"].get(target_currency)
        else:
            print(f"❌ Error fetching exchange rate: {data.get('error-type', 'Unknown error')}")
            return None
    except Exception as e:
        print(f"❌ Request failed: {e}")
        return None


def process_excel_file(target_currency, exchange_rate, excel_file_path, file_extension):
    # Load the Excel file and find the sheet starting with 'sales'
    wb = load_workbook(excel_file_path, keep_vba=True if file_extension == 'xlsm' else False)  # Keep VBA only for .xlsm
    sheet_name = None

    # Check for a sheet that starts with 'sales'
    for sheet in wb.sheetnames:
        if sheet.lower().startswith("sales"):
            sheet_name = sheet
            break

    if not sheet_name:
        print("❌ No sheet found with a name starting with 'sales'.")
        return

    # Select the sheet to process
    ws = wb[sheet_name]
    print(f"Processing sheet: {sheet_name}")

    # Set the header in N1 with the exchange rate
    ws["N1"] = f"Exchange rates EUR/{target_currency} = {exchange_rate}"

    # Process each row, starting from row 2 to skip the header
    for row in range(2, ws.max_row + 1):
        invoice_net = ws.cell(row=row, column=8).value  # Column H: Invoice Net
        invoice_gross = ws.cell(row=row, column=9).value  # Column I: Invoice Gross

        # If values exist for Invoice Net and Invoice Gross, calculate in target currency
        if invoice_net and invoice_gross:
            invoice_net_local = invoice_net * exchange_rate
            invoice_gross_local = invoice_gross * exchange_rate

            # Write converted values into the Local Currency columns (Column K and L)
            ws.cell(row=row, column=11).value = round(invoice_net_local, 2)  # Column K: Local Currency Net
            ws.cell(row=row, column=12).value = round(invoice_gross_local, 2)  # Column L: Local Currency Gross

            # Write the target currency in Column M
            ws.cell(row=row, column=13).value = target_currency  # Column M: Target Currency

    # Save the updated workbook
    if file_extension == "xlsm":
        output_file_path = excel_file_path.replace(".xlsm", f"_{target_currency}.xlsm")  # For macro-enabled files
    else:
        output_file_path = excel_file_path.replace(".xlsx", f"_{target_currency}.xlsx")  # For standard files

    wb.save(output_file_path)
    print(f"✅ Conversion complete. Updated file saved as '{output_file_path}'.")


def get_currency_rate():
    # Load API Key from .env
    load_dotenv()
    API_KEY = os.getenv("ER_API_KEY")

    if not API_KEY:
        print("⚠️  No API key found in .env (ER_API_KEY)")
        API_KEY = input("Please enter your ExchangeRate-API key: ").strip()

    # Prompt user to select the Excel file
    print("Please select the Excel file containing the sales data.")

    Tk().withdraw()  # Hide the root window of tkinter
    excel_file_path = askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xlsm")])  # Support both .xlsx and .xlsm

    if not excel_file_path:
        print("❌ No file selected. Exiting.")
        return

    target_currency = input("Enter target currency (e.g., PLN): ").upper().strip()

    # Get the file extension (xlsx or xlsm)
    file_extension = excel_file_path.split('.')[-1]  # Get the file extension

    # Get the exchange rate for EUR -> target_currency
    exchange_rate = get_exchange_rate("EUR", target_currency, API_KEY)

    if exchange_rate:
        process_excel_file(target_currency, exchange_rate, excel_file_path, file_extension)


if __name__ == "__main__":
    get_currency_rate()
